#!/usr/bin/env python3
import csv
import re
import sys
import os
import json
import time
import argparse
import unicodedata
import urllib.request
import urllib.error
from collections import defaultdict
from pathlib import Path

# --------------------------------------------
# Env loader
# --------------------------------------------

def load_env_file(env_path: "Optional[str]") -> None:
    """Load simple KEY=VALUE lines into os.environ. Ignores comments and blank lines."""
    if not env_path:
        return
    p = Path(env_path)
    if not p.exists():
        return
    try:
        for raw in p.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, val = line.split("=", 1)
            key = key.strip()
            val = val.strip().strip('"').strip("'")
            if key:
                os.environ[key] = val
    except Exception:
        pass

# --------------------------------------------
# Helpers
# --------------------------------------------

def strip_leading_brand(text: str, brand: str = "zummer") -> str:
    if not text:
        return ""
    cleaned = re.sub(rf"^\s*{re.escape(brand)}\b[\s\-_:]*", "", text, flags=re.IGNORECASE)
    return cleaned.strip()


def slugify_for_handle(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize('NFKD', s)
    s = s.encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s)
    s = s.strip('-').lower()
    return s

SIZE_MAP = {
    "xx-small": "xs", "x-small": "xs", "xs": "xs", "extra small": "xs",
    "s": "s", "small": "s",
    "m": "m", "medium": "m",
    "l": "l", "large": "l",
    "xl": "xl", "x-large": "xl",
    "xxl": "2xl", "2xl": "2xl", "xx-large": "2xl",
}

CATEGORY_MAP = {
    "dresses": "Apparel & Accessories > Clothing > Dresses",
    "dress": "Apparel & Accessories > Clothing > Dresses",
    "shirt": "Apparel & Accessories > Clothing > Clothing Tops > Shirts",
    "shirts": "Apparel & Accessories > Clothing > Clothing Tops > Shirts",
    "blouse": "Apparel & Accessories > Clothing > Clothing Tops > Blouses",
    "blouses": "Apparel & Accessories > Clothing > Clothing Tops > Blouses",
    "t-shirt": "Apparel & Accessories > Clothing > Clothing Tops > T-Shirts",
    "t-shirts": "Apparel & Accessories > Clothing > Clothing Tops > T-Shirts",
    "tee": "Apparel & Accessories > Clothing > Clothing Tops > T-Shirts",
    "polo": "Apparel & Accessories > Clothing > Clothing Tops > Polos",
    "polos": "Apparel & Accessories > Clothing > Clothing Tops > Polos",
    "tank top": "Apparel & Accessories > Clothing > Clothing Tops > Tank Tops",
    "tank tops": "Apparel & Accessories > Clothing > Clothing Tops > Tank Tops",
    "sweatshirt": "Apparel & Accessories > Clothing > Clothing Tops > Sweatshirts",
    "sweatshirts": "Apparel & Accessories > Clothing > Clothing Tops > Sweatshirts",
    "cardigan": "Apparel & Accessories > Clothing > Clothing Tops > Cardigans",
    "cardigans": "Apparel & Accessories > Clothing > Clothing Tops > Cardigans",
    "overshirt": "Apparel & Accessories > Clothing > Clothing Tops > Overshirts",
    "overshirts": "Apparel & Accessories > Clothing > Clothing Tops > Overshirts",
    "bodysuit": "Apparel & Accessories > Clothing > Clothing Tops > Bodysuits",
    "bodysuits": "Apparel & Accessories > Clothing > Clothing Tops > Bodysuits",
    "outfit sets": "Apparel & Accessories > Clothing > Outfit Sets",
    "coord": "Apparel & Accessories > Clothing > Outfit Sets",
    "co-ord": "Apparel & Accessories > Clothing > Outfit Sets",
    "co-ords": "Apparel & Accessories > Clothing > Outfit Sets",
    # Bottoms
    "jeans": "Apparel & Accessories > Clothing > Pants > Jeans",
    "jeggings": "Apparel & Accessories > Clothing > Pants > Jeggings",
    "trousers": "Apparel & Accessories > Clothing > Pants > Trousers",
    "cargo pants": "Apparel & Accessories > Clothing > Pants > Cargo Pants",
    "chinos": "Apparel & Accessories > Clothing > Pants > Chinos",
    "joggers": "Apparel & Accessories > Clothing > Pants > Joggers",
    "leggings": "Apparel & Accessories > Clothing > Pants > Leggings",
    "pants": "Apparel & Accessories > Clothing > Pants",
    # Tops umbrella
    "tops": "Apparel & Accessories > Clothing > Clothing Tops",
}

TYPE_MAP = {
    "dresses": "DRESS",
    "dress": "DRESS",
    "shirt": "Shirt",
    "shirts": "Shirt",
    "top": "Top",
    "blouse": "Top",
    "t-shirt": "T-Shirt",
    "t-shirts": "T-Shirt",
    "tee": "T-Shirt",
    "polo": "Polo",
    "polos": "Polo",
    "tank top": "Tank Top",
    "tank tops": "Tank Top",
    "sweatshirt": "Sweatshirt",
    "sweatshirts": "Sweatshirt",
    "cardigan": "Cardigan",
    "cardigans": "Cardigan",
    "overshirt": "Overshirt",
    "overshirts": "Overshirt",
    "bodysuit": "Bodysuit",
    "bodysuits": "Bodysuit",
    "outfit sets": "Co-Ord",
    "coord": "Co-Ord",
    "co-ord": "Co-Ord",
    "co-ords": "Co-Ord",
    # Bottoms
    "jeans": "Jeans",
    "jeggings": "Jeggings",
    "trousers": "Trousers",
    "cargo pants": "Cargo Pants",
    "chinos": "Chinos",
    "joggers": "Joggers",
    "leggings": "Leggings",
    "pants": "Pants",
    # Tops umbrella
    "tops": "Top",
}

ESSENTIAL_HEADERS = [
    "Handle",
    "Title",
    "Body (HTML)",
    "Vendor",
    "Product Category",
    "Type",
    "Tags",
    "Published",
    "Option1 Name",
    "Option1 Value",
    "Variant SKU",
    "Variant Grams",
    "Variant Inventory Tracker",
    "Variant Inventory Qty",
    "Variant Inventory Policy",
    "Variant Fulfillment Service",
    "Variant Price",
    "Variant Compare At Price",
    "Variant Requires Shipping",
    "Variant Taxable",
    "Status",
]

SIZE_ORDER = ["xs", "s", "m", "l", "xl", "2xl"]


def infer_category(article_type: str, title: str) -> str:
    t = (article_type or "").strip().lower()
    if t in CATEGORY_MAP:
        return CATEGORY_MAP[t]
    ttl = (title or "").lower()
    if any(k in ttl for k in ["co-ord", "co ord", "co-ords", "co ords", "coord", "coords", "outfit set", "outfit sets"]):
        return CATEGORY_MAP["co-ord"]
    if "polo" in ttl:
        return CATEGORY_MAP["polo"]
    if "t-shirt" in ttl or "tshirt" in ttl or "tee" in ttl:
        return CATEGORY_MAP["t-shirt"]
    if "shirt" in ttl:
        return CATEGORY_MAP["shirt"]
    if "blouse" in ttl:
        return CATEGORY_MAP["blouse"]
    if "dress" in ttl:
        return CATEGORY_MAP["dress"]
    return "Apparel & Accessories > Clothing"


def infer_type(article_type: str, title: str) -> str:
    t = (article_type or "").strip().lower()
    if t in TYPE_MAP:
        return TYPE_MAP[t]
    ttl = (title or "").lower()
    if any(k in ttl for k in ["co-ord", "co ord", "co-ords", "co ords", "coord", "coords", "outfit set", "outfit sets"]):
        return "Co-Ord"
    if "dress" in ttl:
        return "DRESS"
    if any(k in ttl for k in ["shirt", "t-shirt", "tshirt", "tee", "polo", "blouse", "top"]):
        return "Top"
    return (article_type or "").strip().title() or "Top"


def map_from_source_kind(source_kind: "Optional[str]", fallback_article_type: str, title: str) -> "Tuple[str, str]":
    """Return (product_category, product_type) based on file-of-origin if provided."""
    sk = (source_kind or "").strip().lower()
    if sk in CATEGORY_MAP and sk in TYPE_MAP:
        return CATEGORY_MAP[sk], TYPE_MAP[sk]
    # Some file names might be singular/plural variants
    alias = {
        "shirt": "shirt",
        "shirts": "shirt",
        "top": "tops",
        "tops": "tops",
        "dress": "dress",
        "dresses": "dress",
        "co-ord": "co-ord",
        "co-ords": "co-ords",
        "coord": "co-ord",
        "coords": "co-ords",
        "jean": "jeans",
        "jeans": "jeans",
        "jeggings": "jeggings",
        "trouser": "trousers",
        "trousers": "trousers",
        "pant": "pants",
        "pants": "pants",
    }
    tgt = alias.get(sk)
    if tgt and tgt in CATEGORY_MAP and tgt in TYPE_MAP:
        return CATEGORY_MAP[tgt], TYPE_MAP[tgt]

    # If the source was a generic umbrella (e.g., 'tops' or 'pants'), refine by keywords
    ttl = (title or '').lower()
    art = (fallback_article_type or '').lower()
    def any_in(s: str, keys: list) -> bool:
        return any(k in s for k in keys)

    if sk in ("tops", "top"):
        if any_in(ttl+" "+art, ["t-shirt", "tshirt", "tee"]):
            return CATEGORY_MAP["t-shirt"], TYPE_MAP["t-shirt"]
        if any_in(ttl+" "+art, ["polo"]):
            return CATEGORY_MAP["polo"], TYPE_MAP["polo"]
        if any_in(ttl+" "+art, ["tank", "tank top"]):
            return CATEGORY_MAP["tank top"], TYPE_MAP["tank top"]
        if any_in(ttl+" "+art, ["bodysuit"]):
            return CATEGORY_MAP["bodysuit"], TYPE_MAP["bodysuit"]
        if any_in(ttl+" "+art, ["cardigan"]):
            return CATEGORY_MAP["cardigan"], TYPE_MAP["cardigan"]
        if any_in(ttl+" "+art, ["sweatshirt"]):
            return CATEGORY_MAP["sweatshirt"], TYPE_MAP["sweatshirt"]
        if any_in(ttl+" "+art, ["overshirt"]):
            return CATEGORY_MAP["overshirt"], TYPE_MAP["overshirt"]
        if any_in(ttl+" "+art, ["shirt"]):
            return CATEGORY_MAP["shirt"], TYPE_MAP["shirt"]
        return CATEGORY_MAP["tops"], TYPE_MAP["top"]

    if sk in ("pants",):
        if any_in(ttl+" "+art, ["jean"]):
            return CATEGORY_MAP["jeans"], TYPE_MAP["jeans"]
        if any_in(ttl+" "+art, ["jegging"]):
            return CATEGORY_MAP["jeggings"], TYPE_MAP["jeggings"]
        if any_in(ttl+" "+art, ["trouser"]):
            return CATEGORY_MAP["trousers"], TYPE_MAP["trousers"]
        if any_in(ttl+" "+art, ["cargo"]):
            return CATEGORY_MAP["cargo pants"], TYPE_MAP["cargo pants"]
        if any_in(ttl+" "+art, ["chino"]):
            return CATEGORY_MAP["chinos"], TYPE_MAP["chinos"]
        if any_in(ttl+" "+art, ["jogger"]):
            return CATEGORY_MAP["joggers"], TYPE_MAP["joggers"]
        if any_in(ttl+" "+art, ["legging"]):
            return CATEGORY_MAP["leggings"], TYPE_MAP["leggings"]
        return CATEGORY_MAP["pants"], TYPE_MAP["pants"]
    # Fallback to existing inference
    return infer_category(fallback_article_type, title), infer_type(fallback_article_type, title)


def normalize_size(s: str) -> str:
    if not s:
        return ""
    key = s.strip().lower()
    return SIZE_MAP.get(key, key)


def build_body_html(row: dict) -> str:
    parts = []
    for key in [
        "Product Details",
        "styleNote",
        "materialCareDescription",
        "sizeAndFitDescription",
    ]:
        val = row.get(key) or ""
        val = str(val).strip()
        if val:
            parts.append(strip_leading_brand(val, brand="zummer"))
    if not parts:
        fabric = (row.get("Fabric") or row.get("Fabric 2") or "").strip()
        shape = (row.get("Shape") or "").strip()
        neck = (row.get("Neck") or "").strip()
        sleeve = (row.get("Sleeve Length") or "").strip()
        length = (row.get("Length") or "").strip()
        pattern = (row.get("Pattern") or row.get("Print or Pattern Type") or "").strip()
        attrs = [fabric, shape, neck, sleeve, length, pattern]
        desc = ", ".join([a for a in attrs if a])
        if desc:
            parts.append(desc)
    if not parts:
        return ""
    html = "".join([f"<p>{csv_html_escape(p)}</p>" for p in parts])
    return html

# --------------------------------------------
# LLM integration (optional)
# --------------------------------------------

def build_llm_messages(context: dict) -> list:
    system = (
        "/no_think You are an expert high-quality fashion ecommerce copywriter for Shopify."
        "Write one HTML paragraph (<p>...</p>) containing 3–5 short lines separated by <br>."
        "Keep the tone fresh, playful, conversational, and never too salesy."
        "Do not mention the brand name (Zummer) in your output."
        "Return ONLY the final HTML paragraph of 3-5 lines as the output. Do NOT include any analysis, planning, or explanations other than the final HTML paragraph. Whatever you finally output will go into the Body (HTML) field of the Shopify product and read my real customers."
    )
    attrs = []
    for key in [
        "title", "product_type", "fabric", "shape", "neck", "sleeve_length",
        "length", "pattern", "occasion", "color", "care", "fit", "season", "usage",
        "brand", "audience",
    ]:
        val = context.get(key)
        if val:
            attrs.append(f"{key}: {val}")
    attr_text = "\n".join(attrs)
    user = (
        "Write a 3–5 line Shopify product description (Body HTML) based on the following product details.\n"
        f"{attr_text}\n"
        "Constraints:\n"
        "- Use the Title above as context; do not repeat it verbatim.\n"
        "- Output ONE <p> block only, with each line separated by <br>.\n"
        "- Tone: fresh, playful, conversational (not too salesy).\n"
        "- Focus: fabric, fit, key design details (neckline, sleeve, length).\n"
        "- Include how it suits vacations, brunches, kitty parties, or casual outings.\n"
        "- Make it trendy yet wearable every day.\n"
        "- End with a fun, friendly styling or usage tip as the final line.\n"
        "- Do not start with the brand name.\n"
        "- IMPORTANT: Return ONLY the final <p>...</p> block; no analysis or extra text."
    )
    return [{"role": "system", "content": system}, {"role": "user", "content": user}]


def http_post_json(url: str, payload: dict, api_key: str = "", timeout: int = 30):
    data = json.dumps(payload).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    req = urllib.request.Request(url, data=data, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            body = resp.read()
            return json.loads(body.decode("utf-8"))
    except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, json.JSONDecodeError, Exception):
        return None


def call_openai_chat(base_url: str, api_key: str, model: str, messages: list, temperature: float = 0.7, max_tokens: int = 250, timeout: int = 30) -> str:
    url = base_url.rstrip("/") + "/v1/chat/completions"
    payload = {"model": model, "messages": messages, "temperature": temperature, "max_tokens": max_tokens}
    parsed = http_post_json(url, payload, api_key=api_key, timeout=timeout)
    if not parsed:
        return ""
    return (parsed.get("choices") or [{}])[0].get("message", {}).get("content", "").strip()


def call_openai_completions(base_url: str, api_key: str, model: str, prompt: str, temperature: float = 0.7, max_tokens: int = 250, timeout: int = 30) -> str:
    url = base_url.rstrip("/") + "/v1/completions"
    payload = {"model": model, "prompt": prompt, "temperature": temperature, "max_tokens": max_tokens}
    parsed = http_post_json(url, payload, api_key=api_key, timeout=timeout)
    if not parsed:
        return ""
    choices = parsed.get("choices") or []
    if not choices:
        return ""
    return (choices[0].get("text") or "").strip()


def generate_body_via_llm(handle: str, context: dict, cfg: dict) -> str:
    if not cfg or not cfg.get("enabled"):
        return ""

    cache_dir = cfg.get("cache_dir")
    refresh = bool(cfg.get("refresh"))
    if cache_dir and not refresh:
        cache_path = Path(cache_dir) / f"{handle}.html"
        if cache_path.exists():
            try:
                return cache_path.read_text(encoding="utf-8").strip()
            except Exception:
                pass

    messages = build_llm_messages(context)

    # Resolve API key priority: explicit -> file -> env var (no hardcoded key in repo version)
    api_key = (cfg.get("api_key") or "").strip()
    if not api_key:
        key_file = (cfg.get("api_key_file") or "").strip()
        if key_file:
            try:
                api_key = Path(key_file).read_text(encoding="utf-8").strip()
            except Exception:
                api_key = ""
    if not api_key:
        api_key = os.environ.get(cfg.get("api_key_env", "OPENAI_API_KEY"), "").strip()

    base_url = (cfg.get("base_url") or "https://api.openai.com").strip()
    endpoint = (cfg.get("endpoint") or "chat").strip()  # 'chat' or 'completions'

    # Local endpoints can run without Authorization header
    is_local = base_url.startswith("http://127.0.0.1") or base_url.startswith("http://localhost")
    if not api_key and not is_local:
        return ""  # Cloud requires key

    model = cfg.get("model", "gpt-4o-mini")
    temperature = float(cfg.get("temperature", 0.7))
    max_tokens = int(cfg.get("max_tokens", 250))
    timeout = int(cfg.get("timeout", 30))

    if endpoint == "completions":
        sys_txt = "\n".join([m["content"] for m in messages if m.get("role") == "system"]).strip()
        usr_txt = "\n".join([m["content"] for m in messages if m.get("role") == "user"]).strip()
        prompt = (sys_txt + "\n\n" + usr_txt).strip()
        html = call_openai_completions(base_url, api_key, model, prompt, temperature, max_tokens, timeout)
    else:
        html = call_openai_chat(base_url, api_key, model, messages, temperature, max_tokens, timeout)

    html = (html or "").strip()
    if html:
        lower = html.lower()
        if "<p" in lower:
            try:
                paras = re.findall(r"(?is)<p[^>]*>.*?</p>", html)
                if paras:
                    html = paras[-1].strip()
            except Exception:
                pass
        else:
            html = csv_html_escape(html).replace("\n", "<br>")
            html = f"<p>{html}</p>"

    if html and cache_dir:
        try:
            Path(cache_dir).mkdir(parents=True, exist_ok=True)
            (Path(cache_dir) / f"{handle}.html").write_text(html, encoding="utf-8")
        except Exception:
            pass
    time.sleep(float(cfg.get("rate_sleep", 0)))
    return html


def csv_html_escape(s: str) -> str:
    s = str(s)
    s = s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return s

EXPECTED_HEADER = [
    "styleId","styleGroupId","vendorSkuCode","vendorArticleNumber","vendorArticleName","brand",
    "Manufacturer Name and Address with Pincode","Packer Name and Address with Pincode","Importer Name and Address with Pincode",
    "Country Of Origin","Country Of Origin2","Country Of Origin3","Country Of Origin4","Country Of Origin5",
    "articleType","Brand Size","Standard Size","is Standard Size present on Label","Brand Colour (Remarks)",
    "GTIN","HSN","SKUCode","MRP",
    "AgeGroup","Prominent Colour","Second Prominent Colour","Third Prominent Colour",
    "FashionType","Usage","Year","season",
    "Product Details","styleNote","materialCareDescription","sizeAndFitDescription","productDisplayName","tags","addedDate",
    "Color Variant GroupId","Fabric","Occasion","Shape","Neck","Pattern","Fabric 2","Fabric 3","Length","Sleeve Length",
    "Knit or Woven","Hemline","Print or Pattern Type","Surface Styling","Body Shape ID","Main Trend","Sleeve Styling","Transparency",
    "Fabric Type","Lining","Wash Care","Body or Garment Size","Closure","Add-Ons","Stitch","Character","Sustainable","Number of Pockets",
    "Multipack Set","Number of Items","Contact Brand or Retailer for pre-sales product queries","Where-to-wear","Style Tip","Care for me",
    "Collection Name","Package Contains","BIS Expiry Date","BIS Certificate Image URL","BIS Certificate Number",
    "Bust ( Inches )","Chest ( Inches )","Front Length ( Inches )","Hips ( Inches )","Waist ( Inches )","Across Shoulder ( Inches )",
    "Sleeve-Length ( Inches )","To Fit Bust ( Inches )","To Fit Hip ( Inches )","To Fit Waist ( Inches )",
    "Front Image","Side Image","Back Image","Detail Angle","Look Shot Image","Additional Image 1","Additional Image 2",
]

def read_rows(input_path: Path) -> list:
    """Read a Myntra CSV and return a list of dict rows using the detected header.

    Many Myntra CSVs include a few preface lines (e.g., "Version : 8", section titles)
    before the actual header line. We scan for the header line by locating a row that
    contains the key columns like styleId/vendorSkuCode/etc., then map subsequent rows
    accordingly. This works across different templates (DRESS, JEANS, CO-ORDS, etc.).
    """
    with input_path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = list(csv.reader(f))
    if not reader:
        return []

    header_idx = -1
    header = []
    # Look for a plausible header row anywhere in the file
    for i, row in enumerate(reader):
        lower = [c.strip().lower() for c in row]
        if not lower:
            continue
        has_styleid = any(c == "styleid" for c in lower)
        has_sku = any(c == "vendorskucode" for c in lower)
        has_display = any(c == "productdisplayname" for c in lower)
        has_article = any(c == "articletype" for c in lower)
        if has_styleid and has_sku and (has_display or has_article):
            header_idx = i
            header = [c.strip() for c in row]
            break

    if header_idx == -1:
        # Fallback: try first non-empty row as header
        for i, row in enumerate(reader):
            if any(c.strip() for c in row):
                header_idx = i
                header = [c.strip() for c in row]
                break

    if header_idx == -1:
        return []

    rows = []
    for raw in reader[header_idx + 1 :]:
        if not raw or not any(c.strip() for c in raw):
            continue
        d = {}
        for i, name in enumerate(header):
            if not name:
                continue
            d[name] = raw[i].strip() if i < len(raw) else ""
        # Skip if missing key identity fields entirely
        if not (d.get("styleId") or d.get("styleGroupId") or d.get("SKUCode")):
            continue
        rows.append(d)
    return rows

def read_rows_excel(input_path: Path) -> list:
    """Read a Myntra Excel workbook and return combined list of dict rows from all sheets.

    Uses the same header detection heuristic as read_rows(). Each sheet is scanned
    for a plausible header row; rows are appended with an inferred `_source_kind`
    from the sheet title (lowercased) to aid category/type mapping.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit("openpyxl is required to read Excel files. Install requirements.txt")

    wb = load_workbook(filename=str(input_path), read_only=True, data_only=True)
    all_rows: list = []
    for ws in wb.worksheets:
        # Build a simple list-of-lists from rows (values only)
        sheet_rows = []
        for row in ws.iter_rows(values_only=True):
            sheet_rows.append(["" if v is None else str(v) for v in row])
        if not sheet_rows:
            continue
        # Detect header within the sheet
        header_idx = -1
        header = []
        for i, row in enumerate(sheet_rows):
            lower = [c.strip().lower() for c in row]
            if not lower:
                continue
            has_styleid = any(c == "styleid" for c in lower)
            has_sku = any(c == "vendorskucode" for c in lower)
            has_display = any(c == "productdisplayname" for c in lower)
            has_article = any(c == "articletype" for c in lower)
            if has_styleid and has_sku and (has_display or has_article):
                header_idx = i
                header = [c.strip() for c in row]
                break
        if header_idx == -1:
            for i, row in enumerate(sheet_rows):
                if any(c.strip() for c in row):
                    header_idx = i
                    header = [c.strip() for c in row]
                    break
        if header_idx == -1:
            continue
        # Build dict rows
        for raw in sheet_rows[header_idx + 1 :]:
            if not raw or not any(c.strip() for c in raw):
                continue
            d = {}
            for i, name in enumerate(header):
                if not name:
                    continue
                d[name] = raw[i].strip() if i < len(raw) else ""
            if not (d.get("styleId") or d.get("styleGroupId") or d.get("SKUCode")):
                continue
            d["_source_kind"] = (ws.title or "").strip().lower()
            all_rows.append(d)
    return all_rows

def write_shopify_csv(output_path: Path, rows: list):
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ESSENTIAL_HEADERS)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)

def transform_rows(src_all: list, default_qty: int = 50, default_grams: int = 400, llm_cfg: dict = None, limit_products: int = 0, llm_max_products: int = 0, inventory_qty_blank: bool = False) -> list:
    src = [r for r in src_all if (r.get("vendorSkuCode") or "").strip()]
    if not src:
        return []
    groups = defaultdict(list)
    for row in src:
        key = (row.get("styleGroupId") or row.get("SKUCode") or row.get("styleId") or "").strip()
        groups[key].append(row)

    out_rows = []
    # Apply product limit by style group order if requested
    group_items = list(groups.items())
    if limit_products and limit_products > 0:
        group_items = group_items[:limit_products]

    used_llm_products = 0
    for group_key, items in group_items:
        def size_key(it):
            sz = normalize_size(it.get("Standard Size") or it.get("Brand Size") or "")
            try:
                return SIZE_ORDER.index(sz)
            except ValueError:
                return len(SIZE_ORDER)
        items_sorted = sorted(items, key=size_key)
        items_kept = [it for it in items_sorted if (it.get("vendorSkuCode") or "").strip()]
        if not items_kept:
            continue

        first = items_sorted[0]
        raw_title = (first.get("productDisplayName") or first.get("vendorArticleName") or "").strip()
        title_no_brand = strip_leading_brand(raw_title, brand="zummer")
        title = title_no_brand or raw_title
        handle_base = slugify_for_handle(title) or slugify_for_handle(group_key)
        # Append styleId to handle for uniqueness and traceability
        style_id = (first.get("styleId") or "").strip()
        if style_id:
            handle_base = f"{handle_base}-{slugify_for_handle(style_id)}"

        article_type = (first.get("articleType") or "").strip()
        source_kind = first.get("_source_kind")
        product_category, product_type = map_from_source_kind(source_kind, article_type, raw_title)

        context = {
            "title": title,
            "product_type": product_type,
            "fabric": (first.get("Fabric") or first.get("Fabric 2") or "").strip(),
            "shape": (first.get("Shape") or "").strip(),
            "neck": (first.get("Neck") or "").strip(),
            "sleeve_length": (first.get("Sleeve Length") or "").strip(),
            "length": (first.get("Length") or "").strip(),
            "pattern": (first.get("Pattern") or first.get("Print or Pattern Type") or "").strip(),
            "occasion": (first.get("Occasion") or "").strip(),
            "color": (first.get("Prominent Colour") or "").strip(),
            "care": (first.get("Wash Care") or first.get("materialCareDescription") or "").strip(),
            "fit": (first.get("Fit") or "").strip(),
            "season": (first.get("season") or "").strip(),
            "usage": (first.get("Usage") or "").strip(),
            "brand": (llm_cfg or {}).get("brand") or "Zummer",
            "audience": (llm_cfg or {}).get("audience") or "Modern Indian women, 25–35",
        }
        # Basic non-LLM body from attributes
        body_html = build_body_html(first)

        # Optionally generate description via LLM
        prefer_llm = bool((llm_cfg or {}).get("prefer", False))
        can_use_llm = bool(llm_cfg and llm_cfg.get("enabled"))
        within_llm_cap = (llm_max_products <= 0) or (used_llm_products < llm_max_products)
        if can_use_llm and within_llm_cap and (prefer_llm or not body_html):
            llm_html = generate_body_via_llm(handle_base, context, llm_cfg or {})
            if llm_html:
                body_html = llm_html
                used_llm_products += 1

        selling_price = (first.get("Selling Price") or first.get("Selling price") or "").strip()
        mrp = (first.get("MRP") or "").strip()
        def to_price(v):
            v = re.sub(r"[^0-9\.]+", "", v)
            return v
        price = to_price(selling_price) if selling_price else to_price(mrp)
        # Link MRP to Compare At Price always (when provided)
        compare_at = to_price(mrp) if mrp else ""

        vendor = "Zummer"
        # Keep Variant Inventory Qty empty
        qty_value = ""

        for idx, row in enumerate(items_kept):
            std_size = normalize_size(row.get("Standard Size") or row.get("Brand Size") or "")
            std_size = std_size.upper() if std_size else ""
            sku = (row.get("vendorSkuCode") or "").strip()

            out = {h: "" for h in ESSENTIAL_HEADERS}
            out["Handle"] = handle_base
            out["Title"] = title
            out["Body (HTML)"] = body_html
            out["Vendor"] = vendor
            out["Product Category"] = product_category
            out["Type"] = product_type
            out["Tags"] = ""
            out["Published"] = "true" if idx == 0 else ""
            out["Option1 Name"] = "Size" if idx == 0 else ""
            out["Option1 Value"] = std_size
            out["Variant SKU"] = sku
            out["Variant Grams"] = str(default_grams)
            out["Variant Inventory Tracker"] = "shopify"
            out["Variant Inventory Qty"] = qty_value
            out["Variant Inventory Policy"] = "deny"
            out["Variant Fulfillment Service"] = "manual"
            out["Variant Price"] = price
            out["Variant Compare At Price"] = compare_at
            out["Variant Requires Shipping"] = "true"
            out["Variant Taxable"] = "true"
            out["Status"] = "active" if idx == 0 else ""

            out_rows.append(out)

    return out_rows


def transform(input_path: Path, default_qty: int = 50, default_grams: int = 400, llm_cfg: dict = None, limit_products: int = 0, llm_max_products: int = 0, inventory_qty_blank: bool = False) -> list:
    src_all = read_rows(input_path)
    return transform_rows(
        src_all,
        default_qty=default_qty,
        default_grams=default_grams,
        llm_cfg=llm_cfg,
        limit_products=limit_products,
        llm_max_products=llm_max_products,
        inventory_qty_blank=inventory_qty_blank,
    )


def main():
    # Load .env from project root and CWD first so defaults come from env
    project_env = Path(__file__).resolve().parent.parent / ".env"
    load_env_file(project_env)
    load_env_file(Path.cwd() / ".env")

    # Early parse to pick up --env-file, then load it
    env_only = argparse.ArgumentParser(add_help=False)
    env_only.add_argument("--env-file", default="")
    early_args, remaining = env_only.parse_known_args()
    load_env_file(early_args.env_file or None)

    # Build parser with env-populated defaults
    parser = argparse.ArgumentParser(description="Transform Myntra product CSV to Shopify CSV (variants by size).", parents=[env_only])
    parser.add_argument("--input", help="Path to a single Myntra input CSV")
    parser.add_argument("--input-dir", default="", help="Directory containing multiple Myntra CSV files to combine")
    parser.add_argument("--ignore-dress", action="store_true", help="When using --input-dir, ignore files with 'dress' in the name")
    parser.add_argument("--output", required=True, help="Path to output Shopify CSV")
    parser.add_argument("--default-qty", type=int, default=int(os.getenv("DEFAULT_QTY", "50")), help="Default inventory quantity per variant")
    parser.add_argument("--default-grams", type=int, default=int(os.getenv("DEFAULT_GRAMS", "400")), help="Default weight in grams per variant")
    parser.add_argument("--limit-products", type=int, default=int(os.getenv("LIMIT_PRODUCTS", "0")), help="Limit number of products (style groups) to process; 0 means no limit")
    parser.add_argument("--llm-max-products", type=int, default=int(os.getenv("LLM_MAX_PRODUCTS", "0")), help="Only generate LLM descriptions for the first N products; 0 means unlimited")
    parser.add_argument("--variant-qty-blank", action="store_true", help="Leave Variant Inventory Qty blank in output")
    # LLM options
    parser.add_argument("--llm-enable", action="store_true", help="Enable LLM generation for Body (HTML)")
    parser.add_argument("--llm-provider", default=os.getenv("LLM_PROVIDER", "openai"), help="LLM provider (openai or openai-compatible)")
    parser.add_argument("--llm-base-url", default=os.getenv("LLM_BASE_URL", "https://api.openai.com"), help="Base URL for OpenAI-compatible API (e.g., http://127.0.0.1:1234)")
    parser.add_argument("--llm-endpoint", default=os.getenv("LLM_ENDPOINT", "chat"), choices=["chat","completions"], help="Endpoint to use: chat or completions")
    parser.add_argument("--llm-model", default=os.getenv("LLM_MODEL", "gpt-4o-mini"), help="LLM model name")
    parser.add_argument("--llm-api-key-env", default=os.getenv("LLM_API_KEY_ENV", "OPENAI_API_KEY"), help="Environment variable name holding the API key")
    parser.add_argument("--llm-temperature", type=float, default=float(os.getenv("LLM_TEMPERATURE", "0.7")), help="LLM temperature")
    parser.add_argument("--llm-max-tokens", type=int, default=int(os.getenv("LLM_MAX_TOKENS", "250")), help="LLM max tokens")
    parser.add_argument("--llm-timeout", type=int, default=int(os.getenv("LLM_TIMEOUT", "30")), help="HTTP timeout seconds for LLM calls")
    parser.add_argument("--llm-cache-dir", default=os.getenv("LLM_CACHE_DIR", ""), help="Cache directory for generated Body HTML (by handle)")
    parser.add_argument("--llm-rate-sleep", type=float, default=float(os.getenv("LLM_RATE_SLEEP", "0.0")), help="Sleep seconds between LLM calls")
    parser.add_argument(
        "--llm-prefer",
        action="store_true",
        default=(os.getenv("LLM_PREFER", "").strip().lower() in ("1", "true", "yes")),
        help="Prefer LLM Body (HTML) even when attribute-based description exists",
    )
    parser.add_argument("--llm-brand", default=os.getenv("LLM_BRAND", "Zummer"), help="Brand name to pass into LLM context")
    parser.add_argument("--llm-audience", default=os.getenv("LLM_AUDIENCE", "Modern Indian women, 25–35"), help="Audience description for tone guidance")
    parser.add_argument("--llm-api-key", default=os.getenv("LLM_API_KEY", ""), help="Direct API key string (discouraged; prefer env var)")
    parser.add_argument("--llm-api-key-file", default=os.getenv("LLM_API_KEY_FILE", ""), help="Path to a file containing the API key")
    parser.add_argument("--llm-refresh", action="store_true", default=False, help="Regenerate descriptions even if cache exists")

    args = parser.parse_args(remaining)

    # Collect inputs
    input_rows = []
    input_path = None
    if args.input_dir:
        root = Path(args.input_dir)
        if not root.exists():
            raise FileNotFoundError(f"Input directory not found: {root}")
        files = sorted([p for p in root.glob("*.csv")])
        if args.ignore_dress:
            files = [p for p in files if "dress" not in p.name.lower()]
        for p in files:
            rows = read_rows(p)
            # Tag origin kind from filename to drive category/type mapping later
            fname = p.name.lower()
            kind = ""
            for k in ["co-ords", "co ords", "coord", "coords", "jeans", "jeggings", "trousers", "shirt", "shirts", "top", "tops", "dress", "dresses"]:
                if k.replace(" ", "") in fname.replace(" ", ""):
                    kind = k
                    break
            for r in rows:
                r["_source_kind"] = kind
            if rows:
                input_rows.extend(rows)
    elif args.input:
        input_path = Path(args.input)
        if input_path.suffix.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            rows = read_rows_excel(input_path)
            input_rows.extend(rows)
            input_path = None
    else:
        raise SystemExit("Either --input or --input-dir is required")
    output_path = Path(args.output)

    llm_cfg = None
    if args.llm_enable:
        llm_cfg = {
            "enabled": True,
            "provider": args.llm_provider,
            "base_url": args.llm_base_url,
            "endpoint": args.llm_endpoint,
            "model": args.llm_model,
            "api_key_env": args.llm_api_key_env,
            "api_key": args.llm_api_key,
            "api_key_file": args.llm_api_key_file,
            "temperature": args.llm_temperature,
            "max_tokens": args.llm_max_tokens,
            "timeout": args.llm_timeout,
            "cache_dir": args.llm_cache_dir or None,
            "rate_sleep": args.llm_rate_sleep,
            "brand": args.llm_brand,
            "audience": args.llm_audience,
            "refresh": args.llm_refresh,
            "prefer": bool(args.llm_prefer),
        }

    if input_rows:
        rows = transform_rows(
            input_rows,
            default_qty=args.default_qty,
            default_grams=args.default_grams,
            llm_cfg=llm_cfg,
            limit_products=args.limit_products,
            llm_max_products=args.llm_max_products,
            inventory_qty_blank=args.variant_qty_blank,
        )
    else:
        rows = transform(
            input_path,
            default_qty=args.default_qty,
            default_grams=args.default_grams,
            llm_cfg=llm_cfg,
            limit_products=args.limit_products,
            llm_max_products=args.llm_max_products,
            inventory_qty_blank=args.variant_qty_blank,
        )
    write_shopify_csv(output_path, rows)
    print(f"Wrote {len(rows)} Shopify rows to {output_path}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
