from __future__ import annotations
import csv
from pathlib import Path


def read_rows(input_path: Path) -> list:
    """Read a Myntra CSV and return a list of dict rows using the detected header.

    Tolerates preface lines (e.g., "Version : 8"). Finds a plausible header row
    by locating columns like styleId/vendorSkuCode/productDisplayName.
    """
    with input_path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = list(csv.reader(f))
    if not reader:
        return []

    header_idx = -1
    header: list[str] = []
    for i, row in enumerate(reader):
        lower = [c.strip().lower() for c in row]
        if not lower:
            continue
        has_styleid = any(c == "styleid" for c in lower)
        has_sku = any(c == "vendorskucode" for c in lower)
        has_display = any(c == "productdisplayname" for c in lower)
        has_article = any(c == "articletype" for c in lower)
        if has_styleid and has_sku and (has_display or has_article):
            header_idx = i
            header = [c.strip() for c in row]
            break

    if header_idx == -1:
        for i, row in enumerate(reader):
            if any(c.strip() for c in row):
                header_idx = i
                header = [c.strip() for c in row]
                break

    if header_idx == -1:
        return []

    rows: list[dict] = []
    for raw in reader[header_idx + 1 :]:
        if not raw or not any(c.strip() for c in raw):
            continue
        d: dict = {}
        for i, name in enumerate(header):
            if not name:
                continue
            d[name] = raw[i].strip() if i < len(raw) else ""
        if not (d.get("styleId") or d.get("styleGroupId") or d.get("SKUCode")):
            continue
        rows.append(d)
    return rows


def write_shopify_csv(output_path: Path, rows: list, fieldnames: list[str]) -> None:
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


def _read_rows_xlsx(input_path: Path) -> list:
    def _val_to_str(v):
        # Normalize Excel numeric cells: 5225.0 -> '5225'
        try:
            import numbers
            if isinstance(v, numbers.Number):
                if float(v).is_integer():
                    return str(int(v))
                return str(v)
        except Exception:
            pass
        return "" if v is None else str(v)
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        raise RuntimeError("openpyxl is required to read .xlsx workbooks. Install requirements.txt")
    wb = load_workbook(filename=str(input_path), read_only=True, data_only=True)
    all_rows: list = []
    for ws in wb.worksheets:
        sheet_rows = []
        for row in ws.iter_rows(values_only=True):
            sheet_rows.append([_val_to_str(v) for v in row])
        if not sheet_rows:
            continue
        header_idx = -1
        header = []
        for i, row in enumerate(sheet_rows):
            lower = [c.strip().lower() for c in row]
            if not lower:
                continue
            has_styleid = any(c == "styleid" for c in lower)
            has_sku = any(c == "vendorskucode" for c in lower)
            has_display = any(c == "productdisplayname" for c in lower)
            has_article = any(c == "articletype" for c in lower)
            if has_styleid and has_sku and (has_display or has_article):
                header_idx = i
                header = [c.strip() for c in row]
                break
        if header_idx == -1:
            for i, row in enumerate(sheet_rows):
                if any(c.strip() for c in row):
                    header_idx = i
                    header = [c.strip() for c in row]
                    break
        if header_idx == -1:
            continue
        for raw in sheet_rows[header_idx + 1 :]:
            if not raw or not any(c.strip() for c in raw):
                continue
            d = {}
            for i, name in enumerate(header):
                if not name:
                    continue
                d[name] = raw[i].strip() if i < len(raw) else ""
            if not (d.get("styleId") or d.get("styleGroupId") or d.get("SKUCode")):
                continue
            d["_source_kind"] = (ws.title or "").strip().lower()
            all_rows.append(d)
    return all_rows


def _read_rows_xls(input_path: Path) -> list:
    try:
        import xlrd  # type: ignore
    except Exception:
        raise RuntimeError("xlrd<2.0 is required to read .xls files. Install requirements.txt")
    book = xlrd.open_workbook(str(input_path))
    all_rows: list = []
    for sheet in book.sheets():
        nrows = sheet.nrows
        if nrows == 0:
            continue
        data = []
        for r in range(nrows):
            row = []
            for c in range(sheet.ncols):
                val = sheet.cell_value(r, c)
                # xlrd returns floats for numbers; coerce 5225.0 -> '5225'
                if isinstance(val, float) and float(val).is_integer():
                    row.append(str(int(val)))
                else:
                    row.append("" if val is None else str(val))
            data.append(row)
        header_idx = -1
        header = []
        for i, row in enumerate(data):
            lower = [c.strip().lower() for c in row]
            if not lower:
                continue
            has_styleid = any(c == "styleid" for c in lower)
            has_sku = any(c == "vendorskucode" for c in lower)
            has_display = any(c == "productdisplayname" for c in lower)
            has_article = any(c == "articletype" for c in lower)
            if has_styleid and has_sku and (has_display or has_article):
                header_idx = i
                header = [c.strip() for c in row]
                break
        if header_idx == -1:
            for i, row in enumerate(data):
                if any(c.strip() for c in row):
                    header_idx = i
                    header = [c.strip() for c in row]
                    break
        if header_idx == -1:
            continue
        for raw in data[header_idx + 1 :]:
            if not raw or not any(c.strip() for c in raw):
                continue
            d = {}
            for i, name in enumerate(header):
                if not name:
                    continue
                d[name] = raw[i].strip() if i < len(raw) else ""
            if not (d.get("styleId") or d.get("styleGroupId") or d.get("SKUCode")):
                continue
            d["_source_kind"] = (sheet.name or "").strip().lower()
            all_rows.append(d)
    return all_rows


def read_any_rows(input_path: Path) -> list:
    ext = input_path.suffix.lower()
    if ext == ".csv":
        return read_rows(input_path)
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return _read_rows_xlsx(input_path)
    if ext == ".xls":
        return _read_rows_xls(input_path)
    # default try CSV
    return read_rows(input_path)
